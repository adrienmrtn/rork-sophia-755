#!/usr/bin/env python3
"""Import Histoire courses (4 parts) from Desktop Excel + Glossaire_culture_generale."""

import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from import_courses_and_glossary import (  # noqa: E402
    COURSE_DATA,
    GLOSSARY_DATA,
    SKIP_EXCEL_TITLES,
    collect_glossary_links,
    find_quiz_blocks,
    normalize_content,
    swift_escape,
    swift_string,
    wrap_glossary_terms_in_content,
)

EXCEL_COURS = Path("/Users/didiermartinot/Desktop/Excel_cours_histoire.xlsx")
EXCEL_GLOSS = Path("/Users/didiermartinot/Desktop/Glossaire_culture_generale.xlsx")
SUBJECT_FILTER = "Histoire"
HISTOIRE_SUBJECT = ".histoire"

CLASSIFICATION_MAP = {
    "Référence historique": "referenceHistorique",
    "Concept": "concept",
    "Événement connexe": "evenementConnexe",
    "Événement": "evenementConnexe",
    "Évènement": "evenementConnexe",
    "Personnage": "personnage",
    "Lieu / institution": "lieuInstitution",
    "Institution": "lieuInstitution",
    "Lieu": "lieuInstitution",
    "Mouvement": "concept",
    "Œuvre": "referenceHistorique",
    "Autre": "concept",
    "Date": "referenceHistorique",
    "Loi": "referenceHistorique",
}


def map_classification(label: str) -> str:
    key = CLASSIFICATION_MAP.get(label.strip())
    if not key:
        raise ValueError(f"Unknown classification: {label!r}")
    return key


def load_excel_histoire_courses():
    import openpyxl

    wb = openpyxl.load_workbook(EXCEL_COURS, data_only=True)
    ws = wb["Sheet1"]
    courses = {}
    for r in range(2, ws.max_row + 1):
        matiere = str(ws.cell(r, 2).value or "").strip()
        if matiere != SUBJECT_FILTER:
            continue
        title = str(ws.cell(r, 4).value or "").strip()
        if title in SKIP_EXCEL_TITLES:
            continue
        parts = []
        for part_idx in range(4):
            title_col = 6 + part_idx * 2
            content_col = 7 + part_idx * 2
            part_title = normalize_content(ws.cell(r, title_col).value)
            part_content = normalize_content(ws.cell(r, content_col).value)
            if not part_title and not part_content:
                continue
            parts.append(
                (
                    part_title or f"Partie {part_idx + 1}",
                    part_content,
                )
            )
        courses[title] = {
            "intro": normalize_content(ws.cell(r, 5).value),
            "parts": parts,
        }
    return courses


def load_glossary_for_titles(titles: set[str]):
    import openpyxl

    wb = openpyxl.load_workbook(EXCEL_GLOSS, data_only=True)
    ws = wb["Glossaire"]
    by_course = {}
    for r in range(2, ws.max_row + 1):
        course = str(ws.cell(r, 1).value or "").strip()
        if course not in titles:
            continue
        entry = {
            "classification": str(ws.cell(r, 2).value or "").strip(),
            "term": str(ws.cell(r, 3).value or "").strip(),
            "explanation": normalize_content(ws.cell(r, 4).value),
        }
        if not entry["term"]:
            continue
        by_course.setdefault(course, []).append(entry)
    return by_course


def prepare_excel_course(excel: dict, glossary_rows: list) -> dict:
    def prep(text) -> str:
        normalized = normalize_content(text)
        return wrap_glossary_terms_in_content(normalized, glossary_rows)

    return {
        "intro": prep(excel["intro"]),
        "parts": [(t, prep(c)) for t, c in excel["parts"]],
    }


def lesson_ids_for_course(course_id: str, existing: list[str], part_count: int) -> list[str]:
    intro = next((lid for lid in existing if lid.endswith("_intro")), None)
    if not intro:
        intro = f"{course_id}_intro"
    ids = [intro]
    for i in range(1, part_count + 1):
        suffix = f"_p{i}"
        match = next((lid for lid in existing if lid.endswith(suffix)), None)
        ids.append(match or f"{course_id}{suffix}")
    return ids


def build_lessons(course_id: str, lesson_ids: list[str], excel: dict) -> str:
    lines = [
        f'                LessonPage(id: "{lesson_ids[0]}", title: "Introduction", content: {swift_string(excel["intro"])}),'
    ]
    for (title, content), lid in zip(excel["parts"], lesson_ids[1:]):
        lines.append(
            f'                LessonPage(id: "{lid}", title: {swift_string(title)}, content: {swift_string(content)}),'
        )
    return "\n".join(lines)


def build_course_block(c: dict, excel: dict) -> str:
    part_count = len(excel["parts"])
    lesson_ids = lesson_ids_for_course(c["id"], c["lesson_ids"], part_count)
    lessons = build_lessons(c["id"], lesson_ids, excel)
    quiz_lines = c["quiz_raw"]
    quiz_body = (
        "            quiz: [\n" + quiz_lines + "\n    ]"
        if quiz_lines.strip()
        else "            quiz: []"
    )
    return f"""        Course(
            id: "{c["id"]}",
            title: {swift_string(c["title"])},
            description: {swift_string(c["description"])},
            subject: {c["subject"]},
            subcategory: {swift_string(c["subcategory"])},
            lessons: [
{lessons}
            ],
{quiz_body}
        )"""


def parse_all_courses(text: str):
    pattern = re.compile(
        r"Course\(\s*\n\s*id: \"(course_\d+[^\"]+)\",\s*\n\s*title: \"([^\"]+)\",\s*\n\s*description: \"((?:[^\"\\]|\\.)*)\",\s*\n\s*subject: (\.\w+),\s*\n\s*subcategory: \"((?:[^\"\\]|\\.)*)\",\s*\n\s*lessons: \[",
        re.MULTILINE,
    )
    quiz_blocks = find_quiz_blocks(text)
    matches = list(pattern.finditer(text))
    if len(quiz_blocks) != len(matches):
        raise ValueError(f"Quiz blocks ({len(quiz_blocks)}) != courses ({len(matches)})")

    courses = []
    for idx, m in enumerate(matches):
        cid = m.group(1)
        lessons_start = m.end()
        lessons_end = text.find("\n            ],", lessons_start)
        if lessons_end == -1:
            raise ValueError(f"No lessons end for {cid}")
        lesson_block = text[lessons_start:lessons_end]
        lesson_ids = re.findall(r'LessonPage\(id: "([^"]+)"', lesson_block)
        q_start, q_end, q_content = quiz_blocks[idx]
        courses.append(
            {
                "id": cid,
                "title": m.group(2),
                "description": m.group(3),
                "subject": m.group(4),
                "subcategory": m.group(5),
                "lesson_ids": lesson_ids,
                "quiz_raw": q_content,
            }
        )
    return courses


def parse_glossary_entries(text: str) -> dict:
    pattern = re.compile(
        r'"((?:[^"\\]|\\.)+)\|((?:[^"\\]|\\.)+)": GlossaryEntry\('
        r"displayTerm: \"((?:[^\"\\]|\\.)*)\", "
        r"classification: \.(\w+), "
        r"explanation: \"((?:[^\"\\]|\\.)*)\"\)"
    )
    entries = {}
    for m in pattern.finditer(text):
        course = m.group(1).replace('\\"', '"').replace("\\\\", "\\")
        display = m.group(2).replace('\\"', '"').replace("\\\\", "\\")
        entries[f"{course}|{display}"] = {
            "displayTerm": m.group(3).replace('\\"', '"').replace("\\\\", "\\"),
            "classification": m.group(4),
            "explanation": m.group(5).replace('\\"', '"').replace("\\\\", "\\"),
        }
    return entries


def glossary_links_to_entries(links: dict) -> dict:
    out = {}
    for (course, display), data in links.items():
        cls = map_classification(data["classification"])
        out[f"{course}|{display}"] = {
            "displayTerm": data["displayTerm"],
            "classification": cls,
            "explanation": data["explanation"],
        }
    return out


def entries_from_excel_glossary(glossary_by_course: dict, titles: set[str]) -> dict:
    out = {}
    for course, rows in glossary_by_course.items():
        if course not in titles:
            continue
        for row in rows:
            term = row["term"]
            cls = map_classification(row["classification"])
            key = f"{course}|{term}"
            out[key] = {
                "displayTerm": term,
                "classification": cls,
                "explanation": row["explanation"],
            }
    return out


def write_glossary_swift(entries: dict) -> str:
    lines = []
    for key in sorted(entries.keys()):
        course, display = key.split("|", 1)
        data = entries[key]
        lines.append(
            "        "
            f'"{swift_escape(course)}|{swift_escape(display)}": GlossaryEntry('
            f"displayTerm: {swift_string(data['displayTerm'])}, "
            f"classification: .{data['classification']}, "
            f"explanation: {swift_string(data['explanation'])})"
        )
    body = ",\n".join(lines)
    return f"""import Foundation

/// Glossary entries — merged from imports.
nonisolated enum GlossaryData {{
    static let entries: [String: GlossaryEntry] = [
{body}
    ]
}}
"""


def validate_course_data(text: str) -> None:
    broken = 0
    for line in text.splitlines():
        if "LessonPage" in line and 'content: "' in line:
            r = line.rstrip()
            if not (r.endswith('"),') or r.endswith('")')):
                broken += 1
    if broken:
        raise ValueError(f"CourseData has {broken} multiline content strings")


def main():
    excel_courses = load_excel_histoire_courses()
    print(f"Excel Histoire courses: {len(excel_courses)}")

    course_text = COURSE_DATA.read_text(encoding="utf-8")
    courses = parse_all_courses(course_text)
    histoire_titles = set(excel_courses.keys())

    glossary_rows = load_glossary_for_titles(histoire_titles)
    print(f"Glossary rows for Histoire: {sum(len(v) for v in glossary_rows.values())}")

    prepared = {
        title: prepare_excel_course(excel, glossary_rows.get(title, []))
        for title, excel in excel_courses.items()
    }

    updated = 0
    for c in courses:
        if c["subject"] != HISTOIRE_SUBJECT:
            continue
        if c["title"] not in prepared:
            raise ValueError(f"Missing Excel row for history course: {c['title']}")
        new_block = build_course_block(c, prepared[c["title"]])
        old_pattern = re.compile(
            rf"        Course\(\s*\n\s*id: \"{re.escape(c['id'])}\"[\s\S]*?\n        \)",
            re.MULTILINE,
        )
        if not old_pattern.search(course_text):
            raise ValueError(f"Could not find course block for {c['id']}")
        course_text = old_pattern.sub(lambda _: new_block, course_text, count=1)
        updated += 1

    validate_course_data(course_text)
    COURSE_DATA.write_text(course_text, encoding="utf-8")
    print(f"Updated {updated} history courses in {COURSE_DATA}")

    glossary_text = GLOSSARY_DATA.read_text(encoding="utf-8")
    existing = parse_glossary_entries(glossary_text)
    kept = {k: v for k, v in existing.items() if k.split("|", 1)[0] not in histoire_titles}

    from_excel = entries_from_excel_glossary(glossary_rows, histoire_titles)
    links = collect_glossary_links(prepared, glossary_rows)
    from_content = glossary_links_to_entries(links)

    merged = {**kept, **from_excel, **from_content}
    GLOSSARY_DATA.write_text(write_glossary_swift(merged), encoding="utf-8")
    print(
        f"Glossary: kept {len(kept)} (non-Histoire), "
        f"Histoire {len(merged) - len(kept)}, total {len(merged)}"
    )


if __name__ == "__main__":
    main()
