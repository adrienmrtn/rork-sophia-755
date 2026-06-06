#!/usr/bin/env python3
"""Import course lesson content from Excel_cours_V5.xlsx and generate GlossaryData.swift."""

import re
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
COURSE_DATA = ROOT / "ios/Sophia/Services/CourseData.swift"
GLOSSARY_DATA = ROOT / "ios/Sophia/Services/GlossaryData.swift"
EXCEL_COURS = Path("/Users/didiermartinot/Desktop/Excel_cours_V5.xlsx")
EXCEL_GLOSS = Path("/Users/didiermartinot/Desktop/Glossaire_culture_generale.xlsx")

REMOVE_COURSE_ID = "course_190_persee_et_meduse"
SKIP_EXCEL_TITLES = {"B", ""}

SUBJECT_MAP = {
    "Histoire": ".histoire",
    "Sciences": ".sciences",
    "Littérature": ".litterature",
    "Art": ".art",
    "Mythologie": ".mythologie",
    "Comprendre le monde actuel": ".comprendreLeMonde",
}


def norm_key(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", s.lower())


def swift_escape(s: str) -> str:
    s = s.replace("\\", "\\\\").replace('"', '\\"')
    s = s.replace("\n", "\\n")
    return s


def swift_string(s: str) -> str:
    return f'"{swift_escape(normalize_content(s))}"'


def normalize_content(s) -> str:
    if s is None:
        return ""
    s = str(s)
    s = s.replace("\\n", "\n")
    s = s.replace("/n", "\n")
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()


def load_excel_courses():
    import openpyxl

    wb = openpyxl.load_workbook(EXCEL_COURS, data_only=True)
    ws = wb["Sheet1"]
    courses = {}
    for r in range(2, ws.max_row + 1):
        title = str(ws.cell(r, 4).value or "").strip()
        if title in SKIP_EXCEL_TITLES:
            continue
        courses[title] = {
            "intro": normalize_content(ws.cell(r, 5).value),
            "parts": [
                (
                    normalize_content(ws.cell(r, 6).value),
                    normalize_content(ws.cell(r, 7).value),
                ),
                (
                    normalize_content(ws.cell(r, 8).value),
                    normalize_content(ws.cell(r, 9).value),
                ),
                (
                    normalize_content(ws.cell(r, 10).value),
                    normalize_content(ws.cell(r, 11).value),
                ),
            ],
        }
    return courses


def load_glossary():
    import openpyxl

    wb = openpyxl.load_workbook(EXCEL_GLOSS, data_only=True)
    ws = wb["Glossaire"]
    by_course = {}
    for r in range(2, ws.max_row + 1):
        course = str(ws.cell(r, 1).value or "").strip()
        if course in SKIP_EXCEL_TITLES:
            continue
        entry = {
            "classification": str(ws.cell(r, 2).value or "").strip(),
            "term": str(ws.cell(r, 3).value or "").strip(),
            "explanation": normalize_content(ws.cell(r, 4).value),
        }
        by_course.setdefault(course, []).append(entry)
    return by_course


from typing import Optional


def fuzzy_match_glossary(display_term: str, entries: list) -> Optional[dict]:
    dn = norm_key(display_term)
    if len(dn) < 4:
        return None
    for e in entries:
        tn = norm_key(e["term"])
        if dn == tn:
            return e
        if dn in tn:
            if len(dn) >= 8 or len(dn) / max(len(tn), 1) >= 0.45:
                return e
        if tn in dn:
            if len(tn) >= 8 or len(tn) / max(len(dn), 1) >= 0.45:
                return e
    return None


def fuzzy_match_glossary_extended(display_term: str, entries: list) -> Optional[dict]:
    """Try the display term and common article/plural variants."""
    match = fuzzy_match_glossary(display_term, entries)
    if match:
        return match
    stripped = display_term.strip()
    variants = [stripped]
    if stripped.lower().startswith("le "):
        variants.append(stripped[3:])
    elif stripped.lower().startswith("la "):
        variants.append(stripped[3:])
    elif stripped.lower().startswith("les "):
        variants.append(stripped[4:])
    else:
        variants.extend(
            [
                f"Le {stripped}",
                f"La {stripped}",
                f"Les {stripped}",
                f"L'{stripped}",
            ]
        )
    if stripped.endswith("s") and len(stripped) > 3:
        variants.append(stripped[:-1])
    else:
        variants.append(f"{stripped}s")
    for variant in variants:
        match = fuzzy_match_glossary(variant, entries)
        if match:
            return match
    return None


def wrap_glossary_terms_in_content(content: str, entries: list) -> str:
    """Turn **term** into <Glossary term> when it matches this course's glossary."""
    if not entries:
        return content

    def replacer(match: re.Match[str]) -> str:
        inner = match.group(1).strip()
        if "<" in inner or ">" in inner:
            return match.group(0)
        hit = fuzzy_match_glossary_extended(inner, entries)
        if hit:
            return f"<{hit['term']}>"
        return match.group(0)

    return re.sub(r"\*\*([^*]+)\*\*", replacer, content)


def collect_glossary_links(courses: dict, glossary: dict) -> dict:
    """Map (courseTitle, displayTerm) -> glossary entry for <> terms only."""
    links = {}
    for title, data in courses.items():
        content = data["intro"] + "\n".join(c for _, c in data["parts"])
        entries = glossary.get(title, [])
        for display in re.findall(r"<([^>]+)>", content):
            display = display.strip()
            if not display:
                continue
            key = (title, display)
            if key in links:
                continue
            match = fuzzy_match_glossary_extended(display, entries)
            if match:
                links[key] = {
                    "displayTerm": display,
                    "classification": match["classification"],
                    "explanation": match["explanation"],
                }
    return links


def find_quiz_blocks(text: str):
    blocks = []
    i = 0
    while True:
        start = text.find("            quiz: [", i)
        if start == -1:
            break
        if text.startswith("            quiz: []", start):
            blocks.append((start, start + len("            quiz: []"), ""))
            i = start + 1
            continue
        line_end = text.find("\n", start)
        pos = line_end + 1
        lines = []
        end = None
        while pos < len(text):
            nl = text.find("\n", pos)
            if nl == -1:
                nl = len(text)
            line = text[pos:nl]
            if line.strip() == "]":
                end = nl
                break
            if line.strip().startswith("QuizQuestion("):
                lines.append(line)
            pos = nl + 1
        if end is None:
            raise ValueError(f"Unclosed quiz at {start}")
        blocks.append((start, end, "\n".join(lines)))
        i = start + 1
    return blocks


def parse_courses(text: str):
    pattern = re.compile(
        r"Course\(\s*\n\s*id: \"(course_\d+[^\"]+)\",\s*\n\s*title: \"([^\"]+)\",\s*\n\s*description: \"((?:[^\"\\]|\\.)*)\",\s*\n\s*subject: (\.\w+),\s*\n\s*subcategory: \"((?:[^\"\\]|\\.)*)\",\s*\n\s*lessons: \[",
        re.MULTILINE,
    )
    quiz_blocks = find_quiz_blocks(text)
    matches = list(pattern.finditer(text))
    if len(quiz_blocks) != len(matches):
        raise ValueError(f"Quiz blocks ({len(quiz_blocks)}) != courses ({len(matches)})")

    courses = []
    for idx, m in enumerate(matches):
        cid = m.group(1)
        lessons_start = m.end()
        lessons_end = text.find("\n            ],", lessons_start)
        if lessons_end == -1:
            raise ValueError(f"No lessons end for {cid}")
        lesson_block = text[lessons_start:lessons_end]
        lesson_ids = re.findall(r'LessonPage\(id: "([^"]+)"', lesson_block)
        if len(lesson_ids) != 4:
            raise ValueError(f"{cid} has {len(lesson_ids)} lessons")

        q_start, q_end, q_content = quiz_blocks[idx]
        courses.append(
            {
                "id": cid,
                "title": m.group(2),
                "description": m.group(3),
                "subject": m.group(4),
                "subcategory": m.group(5),
                "lesson_ids": lesson_ids,
                "quiz_raw": q_content,
                "quiz_is_empty": q_content.strip() == "",
            }
        )
    return courses


def build_lessons(course_id: str, lesson_ids: list, excel: dict) -> str:
    intro = excel["intro"]
    lines = [
        f'                LessonPage(id: "{lesson_ids[0]}", title: "Introduction", content: {swift_string(intro)}),'
    ]
    for i, ((title, content), lid) in enumerate(zip(excel["parts"], lesson_ids[1:]), start=1):
        lines.append(
            f'                LessonPage(id: "{lid}", title: {swift_string(title)}, content: {swift_string(content)}),'
        )
    return "\n".join(lines)


def build_course_block(c: dict, excel: dict) -> str:
    lessons = build_lessons(c["id"], c["lesson_ids"], excel)
    quiz_lines = c["quiz_raw"]
    if quiz_lines:
        quiz_body = "            quiz: [\n" + quiz_lines + "\n    ]"
    else:
        quiz_body = "            quiz: []"
    return f"""        Course(
            id: "{c["id"]}",
            title: {swift_string(c["title"])},
            description: {swift_string(c["description"])},
            subject: {c["subject"]},
            subcategory: {swift_string(c["subcategory"])},
            lessons: [
{lessons}
            ],
{quiz_body}
        )"""


def generate_glossary_swift(links: dict) -> str:
    class_map = {
        "Référence historique": "referenceHistorique",
        "Concept": "concept",
        "Événement connexe": "evenementConnexe",
        "Personnage": "personnage",
        "Lieu / institution": "lieuInstitution",
    }
    entries = []
    for (course, display), data in sorted(links.items()):
        cls = class_map.get(data["classification"])
        if not cls:
            raise ValueError(f"Unknown classification: {data['classification']}")
        entries.append(
            "        "
            f'"{swift_escape(course)}|{swift_escape(display)}": GlossaryEntry('
            f"displayTerm: {swift_string(display)}, "
            f"classification: .{cls}, "
            f"explanation: {swift_string(data['explanation'])})"
        )
    body = ",\n".join(entries)
    return f"""import Foundation

/// Auto-generated from Glossaire_culture_generale.xlsx — do not edit manually.
nonisolated enum GlossaryData {{
    static let entries: [String: GlossaryEntry] = [
{body}
    ]
}}
"""


def main():
    excel_courses = load_excel_courses()
    glossary = load_glossary()
    links = collect_glossary_links(excel_courses, glossary)
    print(f"Glossary links for <> terms: {len(links)}")

    text = COURSE_DATA.read_text(encoding="utf-8")
    courses = parse_courses(text)
    courses = [c for c in courses if c["id"] != REMOVE_COURSE_ID]
    print(f"Courses after removal: {len(courses)}")

    blocks = []
    for c in courses:
        if c["title"] not in excel_courses:
            raise ValueError(f"No excel content for {c['title']}")
        blocks.append(build_course_block(c, excel_courses[c["title"]]))

    header = "import Foundation\n\nnonisolated enum CourseData {\n    static let allCourses: [Course] = [\n"
    footer = "\n    ]\n}\n"
    COURSE_DATA.write_text(header + ",\n".join(blocks) + footer, encoding="utf-8")
    GLOSSARY_DATA.write_text(generate_glossary_swift(links), encoding="utf-8")
    print(f"Wrote {COURSE_DATA}")
    print(f"Wrote {GLOSSARY_DATA}")


if __name__ == "__main__":
    main()
