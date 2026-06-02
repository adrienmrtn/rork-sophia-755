#!/usr/bin/env python3
"""Import quiz questions + explanations from Excel_QCM_modifie.xlsx into CourseData.swift."""

import re
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from import_courses_and_glossary import (  # noqa: E402
    COURSE_DATA,
    SKIP_EXCEL_TITLES,
    find_quiz_blocks,
    normalize_content,
    swift_escape,
    swift_string,
)

EXCEL_QCM = Path("/Users/didiermartinot/Desktop/Excel_QCM_modifie.xlsx")
QCM_COUNT = 5
# Column O (15) = QCM1 question; each QCM spans 6 columns.
QCM1_QUESTION_COL = 15


def load_quiz_by_title():
    import openpyxl

    wb = openpyxl.load_workbook(EXCEL_QCM, data_only=True)
    by_title = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for r in range(2, ws.max_row + 1):
            title = str(ws.cell(r, 4).value or "").strip()
            if not title or title in SKIP_EXCEL_TITLES:
                continue
            questions = []
            for q in range(QCM_COUNT):
                base = QCM1_QUESTION_COL + q * 6
                question = normalize_content(ws.cell(r, base).value)
                correct = normalize_content(ws.cell(r, base + 1).value)
                wrong = [
                    normalize_content(ws.cell(r, base + 2).value),
                    normalize_content(ws.cell(r, base + 3).value),
                    normalize_content(ws.cell(r, base + 4).value),
                ]
                explanation = normalize_content(ws.cell(r, base + 5).value)
                if not question and not correct:
                    continue
                options = [correct] + [w for w in wrong if w]
                while len(options) < 4:
                    options.append("")
                options = options[:4]
                questions.append(
                    {
                        "question": question,
                        "options": options,
                        "explanation": explanation,
                    }
                )
            if questions:
                by_title[title] = questions
    return by_title


def parse_all_courses(text: str):
    pattern = re.compile(
        r"Course\(\s*\n\s*id: \"(course_\d+[^\"]+)\",\s*\n\s*title: \"([^\"]+)\",\s*\n\s*description: \"((?:[^\"\\]|\\.)*)\",\s*\n\s*subject: (\.\w+),\s*\n\s*subcategory: \"((?:[^\"\\]|\\.)*)\",\s*\n\s*lessons: \[",
        re.MULTILINE,
    )
    quiz_blocks = find_quiz_blocks(text)
    matches = list(pattern.finditer(text))
    if len(quiz_blocks) != len(matches):
        raise ValueError(f"Quiz blocks ({len(quiz_blocks)}) != courses ({len(matches)})")

    courses = []
    for idx, m in enumerate(matches):
        cid = m.group(1)
        lessons_start = m.end()
        lessons_end = text.find("\n            ],", lessons_start)
        lesson_block = text[lessons_start:lessons_end]
        lesson_ids = re.findall(r'LessonPage\(id: "([^"]+)"', lesson_block)
        quiz_ids = re.findall(
            r'QuizQuestion\(id: "([^"]+)"', quiz_blocks[idx][2]
        )
        courses.append(
            {
                "id": cid,
                "title": m.group(2),
                "lesson_ids": lesson_ids,
                "quiz_ids": quiz_ids,
                "quiz_raw": quiz_blocks[idx][2],
            }
        )
    return courses


def swift_options(options: list[str]) -> str:
    parts = [swift_string(o) for o in options]
    return "[" + ", ".join(parts) + "]"


def format_quiz_question(qid: str, data: dict) -> str:
    return (
        f"        QuizQuestion(id: \"{qid}\", "
        f"question: {swift_string(data['question'])}, "
        f"options: {swift_options(data['options'])}, "
        f"correctIndex: 0, "
        f"explanation: {swift_string(data['explanation'])})"
    )


def build_quiz_block(quiz_ids: list[str], questions: list[dict]) -> str:
    if len(questions) != len(quiz_ids):
        raise ValueError(
            f"Quiz count mismatch: {len(questions)} excel vs {len(quiz_ids)} ids"
        )
    lines = [
        format_quiz_question(qid, q)
        for qid, q in zip(quiz_ids, questions)
    ]
    return "            quiz: [\n" + ",\n".join(lines) + "\n    ]"


def replace_quiz_in_course(text: str, course_id: str, new_quiz_body: str) -> str:
    pattern = re.compile(
        rf"(        Course\(\s*\n\s*id: \"{re.escape(course_id)}\"[\s\S]*?)"
        r"            quiz: \[[\s\S]*?\n    \]",
        re.MULTILINE,
    )

    def replacer(match: re.Match[str]) -> str:
        return match.group(1) + new_quiz_body

    new_text, count = pattern.subn(replacer, text, count=1)
    if count != 1:
        raise ValueError(f"Could not find quiz block for {course_id}")
    return new_text


def main():
    quiz_by_title = load_quiz_by_title()
    print(f"Excel courses with quiz: {len(quiz_by_title)}")

    text = COURSE_DATA.read_text(encoding="utf-8")
    courses = parse_all_courses(text)
    updated = 0
    missing = []

    for c in courses:
        title = c["title"]
        if title not in quiz_by_title:
            missing.append(title)
            continue
        excel_q = quiz_by_title[title]
        if len(excel_q) != len(c["quiz_ids"]):
            raise ValueError(
                f"{title}: excel has {len(excel_q)} questions, app has {len(c['quiz_ids'])}"
            )
        new_quiz = build_quiz_block(c["quiz_ids"], excel_q)
        text = replace_quiz_in_course(text, c["id"], new_quiz)
        updated += 1

    if missing:
        raise ValueError(f"No Excel quiz for {len(missing)} courses: {missing[:5]}")

    COURSE_DATA.write_text(text, encoding="utf-8")
    print(f"Updated quiz for {updated} courses in {COURSE_DATA}")


if __name__ == "__main__":
    main()
